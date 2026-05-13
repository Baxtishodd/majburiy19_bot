import asyncio
import io
from datetime import datetime, timedelta
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from config import ADMIN_IDS
from database import get_pool

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = Router()


# ─── Admin tekshiruvi ─────────────────────────────────────────────────────────

def is_superadmin(user_id: int) -> bool:
    """Faqat .env dagi ADMIN_IDS — superadminlar"""
    return user_id in ADMIN_IDS


async def is_admin(user_id: int) -> bool:
    """ADMIN_IDS yoki DB dagi bot_admins jadvalidagi adminlar"""
    if user_id in ADMIN_IDS:
        return True
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT 1 FROM bot_admins WHERE user_id = %s", (user_id,))
            return bool(await cur.fetchone())


# ─── FSM holatlari ────────────────────────────────────────────────────────────

class BroadcastStates(StatesGroup):
    waiting_post = State()
    waiting_time = State()
    confirm      = State()


class AddAdminStates(StatesGroup):
    waiting_user_id = State()


# ─── Klaviaturalar ────────────────────────────────────────────────────────────

def main_panel_kb(user_id: int):
    b = InlineKeyboardBuilder()
    b.button(text="📊 Statistika",         callback_data="ap:stats")
    b.button(text="👥 Guruhlar",           callback_data="ap:chats")
    b.button(text="📢 Post yuborish",      callback_data="ap:broadcast")
    b.button(text="🕒 Rejalashtirilgan",   callback_data="ap:scheduled")
    b.button(text="📨 Yuborilgan postlar", callback_data="ap:sent_posts:0")
    if is_superadmin(user_id):
        b.button(text="👤 Adminlar",       callback_data="ap:admins")
        b.button(text="📥 Excel hisobot",  callback_data="ap:excel_menu")
    b.adjust(2)
    return b.as_markup()


def back_kb(to="ap:main"):
    b = InlineKeyboardBuilder()
    b.button(text="◀️ Orqaga", callback_data=to)
    return b.as_markup()


def broadcast_target_kb():
    b = InlineKeyboardBuilder()
    b.button(text="👥 Hamma guruhlarga",         callback_data="ap:bc_target:groups")
    b.button(text="📢 Hamma kanallarga",          callback_data="ap:bc_target:channels")
    b.button(text="👤 Hamma foydalanuvchilarga",  callback_data="ap:bc_target:users")
    b.button(text="🌐 Hammaga (barchasi)",        callback_data="ap:bc_target:all")
    b.button(text="◀️ Orqaga",                   callback_data="ap:main")
    b.adjust(1)
    return b.as_markup()


def broadcast_time_kb():
    b = InlineKeyboardBuilder()
    b.button(text="⚡️ Hoziroq yuborish", callback_data="ap:bc_time:now")
    b.button(text="⏰ Vaqt belgilash",    callback_data="ap:bc_time:schedule")
    b.button(text="◀️ Orqaga",           callback_data="ap:broadcast")
    b.adjust(1)
    return b.as_markup()


def confirm_kb():
    b = InlineKeyboardBuilder()
    b.button(text="✅ Tasdiqlash",   callback_data="ap:bc_confirm:yes")
    b.button(text="❌ Bekor qilish", callback_data="ap:bc_confirm:no")
    b.adjust(2)
    return b.as_markup()


def scheduled_list_kb(posts: list):
    b = InlineKeyboardBuilder()
    for post in posts:
        label = f"🗑 {post['scheduled_at'].strftime('%d.%m %H:%M')} — {(post['message_text'] or '')[:20]}..."
        b.button(text=label, callback_data=f"ap:sc_del:{post['id']}")
    b.button(text="◀️ Orqaga", callback_data="ap:main")
    b.adjust(1)
    return b.as_markup()


def admins_panel_kb():
    b = InlineKeyboardBuilder()
    b.button(text="➕ Admin qo'shish", callback_data="ap:admin_add")
    b.button(text="◀️ Orqaga",         callback_data="ap:main")
    b.adjust(1)
    return b.as_markup()


def admin_remove_kb(admins: list):
    b = InlineKeyboardBuilder()
    for a in admins:
        name = a.get("full_name") or a.get("username") or str(a["user_id"])
        b.button(text=f"🗑 {name}", callback_data=f"ap:admin_del:{a['user_id']}")
    b.button(text="➕ Admin qo'shish", callback_data="ap:admin_add")
    b.button(text="◀️ Orqaga",         callback_data="ap:main")
    b.adjust(1)
    return b.as_markup()


def excel_menu_kb():
    b = InlineKeyboardBuilder()
    b.button(text="👤 Foydalanuvchilar",       callback_data="ap:excel:users")
    b.button(text="👥 Guruhlar",               callback_data="ap:excel:groups")
    b.button(text="📢 Kanallar",               callback_data="ap:excel:channels")
    b.button(text="📨 Chop etilgan postlar",   callback_data="ap:excel:posts")
    b.button(text="📅 Oylik hisobot",          callback_data="ap:excel:monthly")
    b.button(text="📦 Hammasi (to'liq ZIP)",   callback_data="ap:excel:all")
    b.button(text="◀️ Orqaga",                callback_data="ap:main")
    b.adjust(2)
    return b.as_markup()


TARGET_LABELS = {
    "groups":   "👥 Hamma guruhlar",
    "channels": "📢 Hamma kanallar",
    "users":    "👤 Hamma foydalanuvchilar",
    "all":      "🌐 Hammaga (barchasi)",
}

SENT_PAGE_SIZE = 5


# ─── /admin ───────────────────────────────────────────────────────────────────

@router.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        await message.answer("❌ Bu buyruq faqat bot adminlari uchun!")
        return
    await state.clear()
    await message.answer(
        "🤖 <b>Admin Panel</b>\n\nQuyidagi bo'limlardan birini tanlang:",
        parse_mode="HTML",
        reply_markup=main_panel_kb(message.from_user.id)
    )


@router.callback_query(F.data == "ap:main")
async def cb_main(call: CallbackQuery, state: FSMContext):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)
    await state.clear()
    await call.message.edit_text(
        "🤖 <b>Admin Panel</b>\n\nQuyidagi bo'limlardan birini tanlang:",
        parse_mode="HTML",
        reply_markup=main_panel_kb(call.from_user.id)
    )


# ─── Statistika ───────────────────────────────────────────────────────────────

@router.callback_query(F.data == "ap:stats")
async def cb_stats(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    await call.message.edit_text("⏳ Statistika yuklanmoqda...", parse_mode="HTML")

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT COUNT(*) FROM bot_chats WHERE chat_type IN ('group','supergroup')")
            groups_count = (await cur.fetchone())[0]

            await cur.execute("SELECT COUNT(*) FROM bot_chats WHERE chat_type = 'channel'")
            channels_count = (await cur.fetchone())[0]

            await cur.execute("SELECT COUNT(*) FROM users")
            users_count = (await cur.fetchone())[0]

            await cur.execute("SELECT COUNT(*) FROM group_referrals")
            refs = (await cur.fetchone())[0]

            await cur.execute("SELECT COUNT(*) FROM scheduled_posts WHERE sent = FALSE")
            pending = (await cur.fetchone())[0]

            await cur.execute("SELECT COUNT(*) FROM scheduled_posts WHERE sent = TRUE")
            sent_total = (await cur.fetchone())[0]

            await cur.execute("SELECT chat_id, chat_type FROM bot_chats")
            chats = await cur.fetchall()

    total_members = 0
    group_members = 0
    channel_members = 0

    for chat_id, chat_type in chats:
        try:
            count = await call.bot.get_chat_member_count(chat_id)
            total_members += count
            if chat_type in ("group", "supergroup"):
                group_members += count
            elif chat_type == "channel":
                channel_members += count
        except Exception:
            pass

    def fmt(n):
        return f"{n:,}".replace(",", " ")

    await call.message.edit_text(
        "📊 <b>Statistika</b>\n\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"👥 Guruhlar soni: <b>{groups_count}</b>\n"
        f"📢 Kanallar soni: <b>{channels_count}</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"👤 Guruhlar auditoriyasi: <b>{fmt(group_members)}</b>\n"
        f"📣 Kanallar auditoriyasi: <b>{fmt(channel_members)}</b>\n"
        f"🌐 Jami auditoriya: <b>{fmt(total_members)}</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"🤝 Bot foydalanuvchilar: <b>{fmt(users_count)}</b>\n"
        f"🔗 Jami referrallar: <b>{fmt(refs)}</b>\n"
        f"🕒 Kutayotgan postlar: <b>{pending}</b>\n"
        f"📨 Yuborilgan postlar: <b>{sent_total}</b>",
        parse_mode="HTML",
        reply_markup=back_kb()
    )


# ─── Guruhlar ro'yxati ────────────────────────────────────────────────────────

@router.callback_query(F.data == "ap:chats")
async def cb_chats(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT chat_id, chat_title, chat_type, username FROM bot_chats ORDER BY added_at DESC LIMIT 30"
            )
            chats = await cur.fetchall()

    if not chats:
        return await call.message.edit_text("📭 Hech qanday guruh/kanal yo'q!", reply_markup=back_kb())

    lines = []
    for chat_id, title, ctype, username in chats:
        icon = "👥" if ctype in ("group", "supergroup") else "📢"
        uname = f" @{username}" if username else ""
        lines.append(f"{icon} <b>{title}</b>{uname}\n   <code>{chat_id}</code>")

    text = "👥 <b>Ulangan guruh va kanallar:</b>\n\n" + "\n\n".join(lines)
    if len(text) > 4000:
        text = text[:4000] + "\n\n..."

    await call.message.edit_text(text, parse_mode="HTML", reply_markup=back_kb())


# ─── Broadcast: maqsad tanlash ───────────────────────────────────────────────

@router.callback_query(F.data == "ap:broadcast")
async def cb_broadcast(call: CallbackQuery, state: FSMContext):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)
    await state.clear()
    await call.message.edit_text(
        "📢 <b>Post yuborish</b>\n\nKimga yuborishni tanlang:",
        parse_mode="HTML",
        reply_markup=broadcast_target_kb()
    )


@router.callback_query(F.data.startswith("ap:bc_target:"))
async def cb_bc_target(call: CallbackQuery, state: FSMContext):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    target = call.data.split(":")[2]
    await state.update_data(target=target)
    await state.set_state(BroadcastStates.waiting_post)

    await call.message.edit_text(
        f"🎯 Maqsad: <b>{TARGET_LABELS[target]}</b>\n\n"
        "✍️ <b>Post matnini yuboring</b>\n\n"
        "Matn, rasm, video yoki hujjat yuborishingiz mumkin.\n"
        "HTML: <b>bold</b>, <i>italic</i>, <code>code</code>, <a href='...'>link</a>\n\n"
        "Bekor qilish: /cancel",
        parse_mode="HTML"
    )


# ─── Broadcast: postni qabul qilish ──────────────────────────────────────────

@router.message(BroadcastStates.waiting_post)
async def bc_receive_post(message: Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        return

    if message.text and message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=main_panel_kb(message.from_user.id))
        return

    if message.photo:
        msg_type, file_id, text = "photo", message.photo[-1].file_id, message.caption or ""
    elif message.video:
        msg_type, file_id, text = "video", message.video.file_id, message.caption or ""
    elif message.document:
        msg_type, file_id, text = "document", message.document.file_id, message.caption or ""
    elif message.animation:
        msg_type, file_id, text = "animation", message.animation.file_id, message.caption or ""
    elif message.text:
        msg_type, file_id, text = "text", None, message.text
    else:
        await message.answer("❌ Bu turdagi xabar qo'llab-quvvatlanmaydi!")
        return

    await state.update_data(msg_type=msg_type, file_id=file_id, text=text)
    await state.set_state(BroadcastStates.waiting_time)

    await message.answer(
        "⏰ <b>Qachon yuborilsin?</b>",
        parse_mode="HTML",
        reply_markup=broadcast_time_kb()
    )


# ─── Broadcast: vaqt ─────────────────────────────────────────────────────────

@router.callback_query(F.data == "ap:bc_time:now")
async def cb_bc_now(call: CallbackQuery, state: FSMContext):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    data = await state.get_data()
    await state.update_data(scheduled_at=None)
    await state.set_state(BroadcastStates.confirm)

    await call.message.edit_text(
        f"📋 <b>Tasdiqlash</b>\n\n"
        f"🎯 Maqsad: <b>{TARGET_LABELS.get(data.get('target', 'all'))}</b>\n"
        f"⚡️ Vaqt: <b>Hoziroq</b>\n\n"
        f"📝 Post:\n{_post_preview(data)}\n\n"
        f"Tasdiqlaysizmi?",
        parse_mode="HTML",
        reply_markup=confirm_kb()
    )


@router.callback_query(F.data == "ap:bc_time:schedule")
async def cb_bc_schedule(call: CallbackQuery, state: FSMContext):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)
    await call.message.edit_text(
        "📅 <b>Vaqt va sanani kiriting</b>\n\n"
        "Format: <code>DD.MM.YYYY HH:MM</code>\n"
        "Misol: <code>25.12.2025 18:30</code>\n\n"
        "Bekor qilish: /cancel",
        parse_mode="HTML"
    )


@router.message(BroadcastStates.waiting_time, F.text)
async def bc_receive_time(message: Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        return

    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=main_panel_kb(message.from_user.id))
        return

    try:
        scheduled_at = datetime.strptime(message.text.strip(), "%d.%m.%Y %H:%M")
        if scheduled_at <= datetime.now():
            await message.answer(
                "❌ Vaqt o'tib ketgan! Kelajak vaqtini kiriting.\n"
                "Format: <code>DD.MM.YYYY HH:MM</code>", parse_mode="HTML"
            )
            return
    except ValueError:
        await message.answer(
            "❌ Noto'g'ri format!\nMisol: <code>25.12.2025 18:30</code>",
            parse_mode="HTML"
        )
        return

    data = await state.get_data()
    await state.update_data(scheduled_at=scheduled_at)
    await state.set_state(BroadcastStates.confirm)

    await message.answer(
        f"📋 <b>Tasdiqlash</b>\n\n"
        f"🎯 Maqsad: <b>{TARGET_LABELS.get(data.get('target', 'all'))}</b>\n"
        f"🕒 Vaqt: <b>{scheduled_at.strftime('%d.%m.%Y %H:%M')}</b>\n\n"
        f"📝 Post:\n{_post_preview(data)}\n\n"
        f"Tasdiqlaysizmi?",
        parse_mode="HTML",
        reply_markup=confirm_kb()
    )


# ─── Broadcast: tasdiqlash ────────────────────────────────────────────────────

@router.callback_query(F.data == "ap:bc_confirm:no")
async def cb_bc_no(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("❌ Bekor qilindi.", reply_markup=back_kb())


@router.callback_query(F.data == "ap:bc_confirm:yes")
async def cb_bc_yes(call: CallbackQuery, state: FSMContext):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    data = await state.get_data()
    await state.clear()

    scheduled_at = data.get("scheduled_at") or datetime.now()

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """INSERT INTO scheduled_posts
                   (admin_id, message_text, message_type, file_id, target_type, scheduled_at)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (call.from_user.id, data.get("text", ""), data.get("msg_type", "text"),
                 data.get("file_id"), data.get("target", "all"), scheduled_at)
            )
            post_id = cur.lastrowid

    if scheduled_at <= datetime.now() + timedelta(seconds=5):
        await call.message.edit_text("⏳ Post yuborilmoqda...")
        sent, fail, audience = await _do_broadcast(call.bot, post_id, data)
        await call.message.edit_text(
            f"✅ <b>Post yuborildi!</b>\n\n"
            f"✔️ Muvaffaqiyatli: <b>{sent}</b>\n"
            f"❌ Xatolik: <b>{fail}</b>\n"
            f"👁 Jami auditoriya: <b>{audience:,}</b>".replace(",", " "),
            parse_mode="HTML",
            reply_markup=back_kb()
        )
    else:
        await call.message.edit_text(
            f"✅ <b>Post rejalashtirildi!</b>\n\n"
            f"🕒 Yuborilish vaqti: <b>{scheduled_at.strftime('%d.%m.%Y %H:%M')}</b>",
            parse_mode="HTML",
            reply_markup=back_kb()
        )


# ─── Rejalashtirilgan postlar ─────────────────────────────────────────────────

@router.callback_query(F.data == "ap:scheduled")
async def cb_scheduled(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """SELECT id, message_text, message_type, scheduled_at, target_type
                   FROM scheduled_posts WHERE sent = FALSE ORDER BY scheduled_at ASC"""
            )
            rows = await cur.fetchall()

    if not rows:
        return await call.message.edit_text("📭 Rejalashtirilgan postlar yo'q.", reply_markup=back_kb())

    posts = [{"id": r[0], "message_text": r[1], "message_type": r[2],
               "scheduled_at": r[3], "target_type": r[4]} for r in rows]

    lines = []
    for p in posts:
        preview = (p["message_text"] or f"[{p['message_type']}]")[:25]
        target_label = TARGET_LABELS.get(p["target_type"], p["target_type"])
        lines.append(
            f"🕒 <b>{p['scheduled_at'].strftime('%d.%m.%Y %H:%M')}</b>\n"
            f"   🎯 {target_label} — {preview}..."
        )

    await call.message.edit_text(
        "🕒 <b>Rejalashtirilgan postlar:</b>\n\n" + "\n\n".join(lines) +
        "\n\n<i>O'chirish uchun tugmani bosing:</i>",
        parse_mode="HTML",
        reply_markup=scheduled_list_kb(posts)
    )


@router.callback_query(F.data.startswith("ap:sc_del:"))
async def cb_sc_delete(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    post_id = int(call.data.split(":")[2])
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("DELETE FROM scheduled_posts WHERE id = %s AND sent = FALSE", (post_id,))

    await call.answer("🗑 Post o'chirildi!", show_alert=True)
    await cb_scheduled(call)


# ─── Yuborilgan postlar ro'yxati ──────────────────────────────────────────────

@router.callback_query(F.data.startswith("ap:sent_posts:"))
async def cb_sent_posts(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    page = int(call.data.split(":")[2])

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT COUNT(*) FROM scheduled_posts WHERE sent = TRUE")
            total = (await cur.fetchone())[0]

            await cur.execute(
                """SELECT id, message_text, message_type, target_type, scheduled_at
                   FROM scheduled_posts
                   WHERE sent = TRUE
                   ORDER BY scheduled_at DESC
                   LIMIT %s OFFSET %s""",
                (SENT_PAGE_SIZE, page * SENT_PAGE_SIZE)
            )
            rows = await cur.fetchall()

    if not rows:
        return await call.message.edit_text("📭 Hali hech qanday post yuborilmagan.", reply_markup=back_kb())

    total_pages = max(1, (total + SENT_PAGE_SIZE - 1) // SENT_PAGE_SIZE)

    b = InlineKeyboardBuilder()

    # Har bir post uchun bosiladigan tugma
    for pid, text, mtype, target, sched_at in rows:
        icon = {"text": "📝", "photo": "🖼", "video": "🎬",
                "document": "📄", "animation": "🎞"}.get(mtype, "📝")
        preview = (text or f"[{mtype}]")[:28]
        label = f"{icon} {sched_at.strftime('%d.%m %H:%M')} — {preview}"
        b.button(text=label, callback_data=f"ap:sent_detail:{pid}:{page}")

    # Sahifalash tugmalari
    nav_row = []
    if page > 0:
        nav_row.append(("◀️ Oldingi", f"ap:sent_posts:{page - 1}"))
    if page < total_pages - 1:
        nav_row.append(("Keyingi ▶️", f"ap:sent_posts:{page + 1}"))
    for label, cb in nav_row:
        b.button(text=label, callback_data=cb)
    b.button(text="◀️ Orqaga", callback_data="ap:main")

    # Har bir post o'z qatorida, nav tugmalari yonma-yon
    b.adjust(*([1] * len(rows)), len(nav_row) if nav_row else 1, 1)

    await call.message.edit_text(
        f"📨 <b>Yuborilgan postlar</b>  [{page + 1}/{total_pages}]\n\n"
        f"<i>Ko'rish uchun postga bosing 👇</i>",
        parse_mode="HTML",
        reply_markup=b.as_markup()
    )


# ─── Yuborilgan post detail sahifasi ─────────────────────────────────────────

@router.callback_query(F.data.startswith("ap:sent_detail:"))
async def cb_sent_detail(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    parts = call.data.split(":")
    post_id = int(parts[2])
    back_page = int(parts[3]) if len(parts) > 3 else 0

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                """SELECT id, message_text, message_type, file_id, target_type,
                          sent_count, fail_count, total_audience, scheduled_at, admin_id
                   FROM scheduled_posts WHERE id = %s""",
                (post_id,)
            )
            row = await cur.fetchone()

    if not row:
        return await call.answer("❌ Post topilmadi!", show_alert=True)

    pid, msg_text, mtype, file_id, target, sent_c, fail_c, audience, sched_at, admin_id = row

    def fmt(n):
        return f"{n:,}".replace(",", " ")

    type_icons = {"text": "📝 Matn", "photo": "🖼 Rasm", "video": "🎬 Video",
                  "document": "📄 Fayl", "animation": "🎞 GIF"}
    type_label = type_icons.get(mtype, f"📎 {mtype}")
    target_label = TARGET_LABELS.get(target, target)

    # Matn previewni to'liq ko'rsatish (max 800 belgi)
    content_preview = ""
    if msg_text:
        if len(msg_text) > 800:
            content_preview = f"\n\n📄 <b>Matn (qisqartirilgan):</b>\n{msg_text[:800]}..."
        else:
            content_preview = f"\n\n📄 <b>Matn:</b>\n{msg_text}"

    detail_text = (
        f"📨 <b>Post #{pid} tafsilotlari</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"📁 Turi: <b>{type_label}</b>\n"
        f"🎯 Maqsad: <b>{target_label}</b>\n"
        f"📅 Yuborilgan: <b>{sched_at.strftime('%d.%m.%Y %H:%M')}</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"✔️ Muvaffaqiyatli: <b>{fmt(sent_c)}</b>\n"
        f"❌ Xatolik: <b>{fmt(fail_c)}</b>\n"
        f"👁 Jami auditoriya: <b>{fmt(audience)}</b>"
        f"{content_preview}"
    )

    b = InlineKeyboardBuilder()

    # Agar media bo'lsa, previewni ko'rish uchun tugma
    if mtype != "text" and file_id:
        b.button(text="🖼 Mediani ko'rish", callback_data=f"ap:sent_media:{post_id}:{back_page}")

    b.button(text="◀️ Ro'yxatga qaytish", callback_data=f"ap:sent_posts:{back_page}")
    b.button(text="🏠 Bosh sahifa", callback_data="ap:main")
    b.adjust(1)

    await call.message.edit_text(detail_text, parse_mode="HTML", reply_markup=b.as_markup())


# ─── Post mediasini ko'rish ───────────────────────────────────────────────────

@router.callback_query(F.data.startswith("ap:sent_media:"))
async def cb_sent_media(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("❌ Ruxsat yo'q!", show_alert=True)

    parts = call.data.split(":")
    post_id = int(parts[2])
    back_page = int(parts[3]) if len(parts) > 3 else 0

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT message_type, file_id, message_text FROM scheduled_posts WHERE id = %s",
                (post_id,)
            )
            row = await cur.fetchone()

    if not row or not row[1]:
        return await call.answer("❌ Media topilmadi!", show_alert=True)

    mtype, file_id, caption = row

    b = InlineKeyboardBuilder()
    b.button(text="◀️ Orqaga", callback_data=f"ap:sent_detail:{post_id}:{back_page}")

    try:
        if mtype == "photo":
            await call.message.answer_photo(file_id, caption=caption or "", parse_mode="HTML", reply_markup=b.as_markup())
        elif mtype == "video":
            await call.message.answer_video(file_id, caption=caption or "", parse_mode="HTML", reply_markup=b.as_markup())
        elif mtype == "document":
            await call.message.answer_document(file_id, caption=caption or "", parse_mode="HTML", reply_markup=b.as_markup())
        elif mtype == "animation":
            await call.message.answer_animation(file_id, caption=caption or "", parse_mode="HTML", reply_markup=b.as_markup())
        await call.answer()
    except Exception as e:
        await call.answer(f"❌ Xatolik: {str(e)[:50]}", show_alert=True)


# ─── Adminlar boshqaruvi (faqat superadminlar) ───────────────────────────────

@router.callback_query(F.data == "ap:admins")
async def cb_admins(call: CallbackQuery):
    if not is_superadmin(call.from_user.id):
        return await call.answer("❌ Faqat superadminlar uchun!", show_alert=True)

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT user_id, username, full_name, added_at FROM bot_admins ORDER BY added_at DESC"
            )
            rows = await cur.fetchall()

    sa_lines = []
    for uid in ADMIN_IDS:
        sa_lines.append(f"⭐️ <code>{uid}</code> — superadmin (.env)")

    db_admins = []
    db_lines = []
    for user_id, username, full_name, added_at in rows:
        name = full_name or username or str(user_id)
        uname = f" @{username}" if username else ""
        db_lines.append(f"👤 <b>{name}</b>{uname} — <code>{user_id}</code>\n   📅 {added_at.strftime('%d.%m.%Y')}")
        db_admins.append({"user_id": user_id, "full_name": full_name, "username": username})

    text = "👤 <b>Adminlar ro'yxati</b>\n\n"
    text += "<b>Superadminlar:</b>\n" + "\n".join(sa_lines) + "\n\n"
    if db_lines:
        text += "<b>Qo'shilgan adminlar:</b>\n" + "\n\n".join(db_lines)
    else:
        text += "<i>Qo'shimcha admin yo'q.</i>"

    await call.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=admin_remove_kb(db_admins)
    )


@router.callback_query(F.data == "ap:admin_add")
async def cb_admin_add(call: CallbackQuery, state: FSMContext):
    if not is_superadmin(call.from_user.id):
        return await call.answer("❌ Faqat superadminlar uchun!", show_alert=True)

    await state.set_state(AddAdminStates.waiting_user_id)
    await call.message.edit_text(
        "➕ <b>Yangi admin qo'shish</b>\n\n"
        "Adminning Telegram <b>User ID</b> sini yuboring:\n"
        "<i>(Foydalanuvchi botga /start bosishi kerak)</i>\n\n"
        "Bekor qilish: /cancel",
        parse_mode="HTML"
    )


@router.message(AddAdminStates.waiting_user_id)
async def admin_receive_id(message: Message, state: FSMContext):
    if not is_superadmin(message.from_user.id):
        return

    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=main_panel_kb(message.from_user.id))
        return

    text = message.text.strip() if message.text else ""
    if not text.lstrip("-").isdigit():
        await message.answer("❌ Faqat raqam kiriting (User ID)!\nMisol: <code>123456789</code>", parse_mode="HTML")
        return

    new_admin_id = int(text)

    if is_superadmin(new_admin_id):
        await message.answer("ℹ️ Bu foydalanuvchi allaqachon superadmin!")
        return

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT username, full_name FROM users WHERE id = %s", (new_admin_id,))
            user_row = await cur.fetchone()

            if not user_row:
                await message.answer(
                    "❌ Bu ID dagi foydalanuvchi botda topilmadi.\n"
                    "Avval foydalanuvchi /start bosishi kerak!",
                    parse_mode="HTML"
                )
                return

            username, full_name = user_row

            await cur.execute("SELECT 1 FROM bot_admins WHERE user_id = %s", (new_admin_id,))
            if await cur.fetchone():
                await message.answer("ℹ️ Bu foydalanuvchi allaqachon admin!")
                return

            await cur.execute(
                "INSERT INTO bot_admins (user_id, username, full_name, added_by) VALUES (%s, %s, %s, %s)",
                (new_admin_id, username, full_name, message.from_user.id)
            )

    await state.clear()
    name = full_name or username or str(new_admin_id)
    await message.answer(
        f"✅ <b>{name}</b> admin sifatida qo'shildi!\n"
        f"🆔 <code>{new_admin_id}</code>",
        parse_mode="HTML",
        reply_markup=main_panel_kb(message.from_user.id)
    )

    try:
        await message.bot.send_message(
            new_admin_id,
            "🎉 <b>Tabriklaymiz!</b>\n\nSiz bot admini sifatida tayinlandingiz.\n/admin buyrug'ini bosing.",
            parse_mode="HTML"
        )
    except Exception:
        pass


@router.callback_query(F.data.startswith("ap:admin_del:"))
async def cb_admin_del(call: CallbackQuery):
    if not is_superadmin(call.from_user.id):
        return await call.answer("❌ Faqat superadminlar uchun!", show_alert=True)

    del_id = int(call.data.split(":")[2])
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("SELECT full_name, username FROM bot_admins WHERE user_id = %s", (del_id,))
            row = await cur.fetchone()
            await cur.execute("DELETE FROM bot_admins WHERE user_id = %s", (del_id,))

    name = (row[0] or row[1] or str(del_id)) if row else str(del_id)
    await call.answer(f"🗑 {name} adminlikdan olib tashlandi!", show_alert=True)
    await cb_admins(call)


# ─── Excel eksport (faqat superadminlar) ────────────────────────────────────

def _excel_style_header(ws, headers: list, header_fill="1F4E79"):
    """Ustun sarlavhalarini formatlaydi"""
    fill = PatternFill("solid", start_color=header_fill)
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 22


def _excel_style_rows(ws, start_row=2):
    """Qatorlarni alternativ rang bilan bezaydi"""
    border_side = Side(style="thin", color="D0D0D0")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    fills = [
        PatternFill("solid", start_color="F2F7FF"),
        PatternFill("solid", start_color="FFFFFF"),
    ]
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), 0):
        for cell in row:
            cell.fill = fills[row_idx % 2]
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _excel_auto_width(ws):
    """Ustun kengliklarini avtomatik moslashtiradi"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def _wb_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _build_users_sheet(wb, pool):
    ws = wb.create_sheet("Foydalanuvchilar")
    headers = ["#", "ID", "Username", "To'liq ism", "Referrer ID",
               "Referral soni", "Balans", "Ro'yxatdan o'tgan sana"]
    _excel_style_header(ws, headers)

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT id, username, full_name, referrer_id, referral_count, balance, created_at "
                "FROM users ORDER BY created_at DESC"
            )
            rows = await cur.fetchall()

    for i, (uid, uname, fname, ref_id, ref_cnt, bal, created) in enumerate(rows, 1):
        ws.append([
            i, uid,
            f"@{uname}" if uname else "",
            fname or "",
            ref_id or "",
            ref_cnt or 0,
            bal or 0,
            created.strftime("%d.%m.%Y %H:%M") if created else "",
        ])

    _excel_style_rows(ws)
    _excel_auto_width(ws)
    return len(rows)


async def _build_groups_sheet(wb, pool):
    ws = wb.create_sheet("Guruhlar")
    headers = ["#", "Chat ID", "Nomi", "Username", "Qo'shilgan sana"]
    _excel_style_header(ws, headers)

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT chat_id, chat_title, username, added_at FROM bot_chats "
                "WHERE chat_type IN ('group','supergroup') ORDER BY added_at DESC"
            )
            rows = await cur.fetchall()

    for i, (cid, title, uname, added) in enumerate(rows, 1):
        ws.append([
            i, cid, title or "",
            f"@{uname}" if uname else "",
            added.strftime("%d.%m.%Y %H:%M") if added else "",
        ])

    _excel_style_rows(ws)
    _excel_auto_width(ws)
    return len(rows)


async def _build_channels_sheet(wb, pool):
    ws = wb.create_sheet("Kanallar")
    headers = ["#", "Chat ID", "Nomi", "Username", "Qo'shilgan sana"]
    _excel_style_header(ws, headers)

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT chat_id, chat_title, username, added_at FROM bot_chats "
                "WHERE chat_type = 'channel' ORDER BY added_at DESC"
            )
            rows = await cur.fetchall()

    for i, (cid, title, uname, added) in enumerate(rows, 1):
        ws.append([
            i, cid, title or "",
            f"@{uname}" if uname else "",
            added.strftime("%d.%m.%Y %H:%M") if added else "",
        ])

    _excel_style_rows(ws)
    _excel_auto_width(ws)
    return len(rows)


async def _build_posts_sheet(wb, pool):
    ws = wb.create_sheet("Chop etilgan postlar")
    headers = ["#", "ID", "Admin ID", "Turi", "Maqsad", "Matn (qisqa)",
               "Yuborilgan", "Xato", "Auditoriya", "Yuborilgan vaqt"]
    _excel_style_header(ws, headers)

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT id, admin_id, message_type, target_type, message_text, "
                "sent_count, fail_count, total_audience, scheduled_at "
                "FROM scheduled_posts WHERE sent = TRUE ORDER BY scheduled_at DESC"
            )
            rows = await cur.fetchall()

    target_map = {"groups": "Guruhlar", "channels": "Kanallar",
                  "users": "Foydalanuvchilar", "all": "Hammasi"}

    for i, (pid, admin_id, mtype, target, text, sent_c, fail_c, audience, sched) in enumerate(rows, 1):
        preview = (text or "")[:60] + ("..." if text and len(text) > 60 else "")
        ws.append([
            i, pid, admin_id,
            mtype or "text",
            target_map.get(target, target or ""),
            preview,
            sent_c or 0, fail_c or 0, audience or 0,
            sched.strftime("%d.%m.%Y %H:%M") if sched else "",
        ])

    _excel_style_rows(ws)
    _excel_auto_width(ws)
    return len(rows)


async def _build_monthly_sheet(wb, pool):
    ws = wb.create_sheet("Oylik hisobot")
    headers = ["Oy", "Yangi foydalanuvchilar", "Yuborilgan postlar",
               "Jami xabar yuborildi", "Jami xato", "Jami auditoriya"]
    _excel_style_header(ws, headers, header_fill="145A32")

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            # Oylik foydalanuvchilar
            await cur.execute(
                "SELECT DATE_FORMAT(created_at, '%Y-%m') as mon, COUNT(*) "
                "FROM users GROUP BY mon ORDER BY mon DESC LIMIT 24"
            )
            user_rows = {r[0]: r[1] for r in await cur.fetchall()}

            # Oylik postlar
            await cur.execute(
                "SELECT DATE_FORMAT(scheduled_at, '%Y-%m') as mon, "
                "COUNT(*), SUM(sent_count), SUM(fail_count), SUM(total_audience) "
                "FROM scheduled_posts WHERE sent = TRUE "
                "GROUP BY mon ORDER BY mon DESC LIMIT 24"
            )
            post_rows = {r[0]: (r[1], r[2] or 0, r[3] or 0, r[4] or 0) for r in await cur.fetchall()}

    # Barcha oylarni birlashtirish
    all_months = sorted(set(list(user_rows.keys()) + list(post_rows.keys())), reverse=True)

    for mon in all_months:
        try:
            dt = datetime.strptime(mon, "%Y-%m")
            mon_label = dt.strftime("%B %Y")  # "January 2025"
        except Exception:
            mon_label = mon

        u_cnt = user_rows.get(mon, 0)
        p_data = post_rows.get(mon, (0, 0, 0, 0))
        ws.append([mon_label, u_cnt, p_data[0], p_data[1], p_data[2], p_data[3]])

    _excel_style_rows(ws)
    _excel_auto_width(ws)
    return len(all_months)


@router.callback_query(F.data == "ap:excel_menu")
async def cb_excel_menu(call: CallbackQuery):
    if not is_superadmin(call.from_user.id):
        return await call.answer("❌ Faqat superadminlar uchun!", show_alert=True)

    await call.message.edit_text(
        "📥 <b>Excel hisobot yuklab olish</b>\n\n"
        "Qaysi ma'lumotlarni yuklab olishni tanlang:",
        parse_mode="HTML",
        reply_markup=excel_menu_kb()
    )


@router.callback_query(F.data.startswith("ap:excel:"))
async def cb_excel_export(call: CallbackQuery):
    if not is_superadmin(call.from_user.id):
        return await call.answer("❌ Faqat superadminlar uchun!", show_alert=True)

    if not OPENPYXL_AVAILABLE:
        return await call.answer(
            "❌ openpyxl kutubxonasi o'rnatilmagan!\n"
            "pip install openpyxl", show_alert=True
        )

    export_type = call.data.split(":")[2]
    await call.message.edit_text("⏳ Excel fayl tayyorlanmoqda...", parse_mode="HTML")

    pool = await get_pool()
    now_str = datetime.now().strftime("%d.%m.%Y_%H-%M")

    try:
        wb = openpyxl.Workbook()
        # Default sheet ni o'chirish
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        counts = {}

        if export_type == "users":
            counts["users"] = await _build_users_sheet(wb, pool)
            filename = f"Foydalanuvchilar_{now_str}.xlsx"
            caption = f"👤 <b>Foydalanuvchilar ro'yxati</b>\n📊 Jami: <b>{counts['users']} ta</b>"

        elif export_type == "groups":
            counts["groups"] = await _build_groups_sheet(wb, pool)
            filename = f"Guruhlar_{now_str}.xlsx"
            caption = f"👥 <b>Guruhlar ro'yxati</b>\n📊 Jami: <b>{counts['groups']} ta</b>"

        elif export_type == "channels":
            counts["channels"] = await _build_channels_sheet(wb, pool)
            filename = f"Kanallar_{now_str}.xlsx"
            caption = f"📢 <b>Kanallar ro'yxati</b>\n📊 Jami: <b>{counts['channels']} ta</b>"

        elif export_type == "posts":
            counts["posts"] = await _build_posts_sheet(wb, pool)
            filename = f"Postlar_{now_str}.xlsx"
            caption = f"📨 <b>Yuborilgan postlar</b>\n📊 Jami: <b>{counts['posts']} ta</b>"

        elif export_type == "monthly":
            counts["months"] = await _build_monthly_sheet(wb, pool)
            filename = f"Oylik_hisobot_{now_str}.xlsx"
            caption = f"📅 <b>Oylik statistika hisoboti</b>\n📊 Jami: <b>{counts['months']} oy</b>"

        elif export_type == "all":
            counts["users"]    = await _build_users_sheet(wb, pool)
            counts["groups"]   = await _build_groups_sheet(wb, pool)
            counts["channels"] = await _build_channels_sheet(wb, pool)
            counts["posts"]    = await _build_posts_sheet(wb, pool)
            counts["months"]   = await _build_monthly_sheet(wb, pool)
            filename = f"To'liq_hisobot_{now_str}.xlsx"
            caption = (
                f"📦 <b>To'liq hisobot</b>\n\n"
                f"👤 Foydalanuvchilar: <b>{counts['users']} ta</b>\n"
                f"👥 Guruhlar: <b>{counts['groups']} ta</b>\n"
                f"📢 Kanallar: <b>{counts['channels']} ta</b>\n"
                f"📨 Postlar: <b>{counts['posts']} ta</b>\n"
                f"📅 Oylik hisobot: <b>{counts['months']} oy</b>"
            )
        else:
            await call.message.edit_text("❌ Noto'g'ri tur!", reply_markup=back_kb("ap:excel_menu"))
            return

        file_bytes = _wb_to_bytes(wb)
        excel_file = BufferedInputFile(file_bytes, filename=filename)

        await call.message.delete()
        await call.bot.send_document(
            call.from_user.id,
            document=excel_file,
            caption=caption,
            parse_mode="HTML",
            reply_markup=back_kb("ap:excel_menu")
        )

    except Exception as e:
        await call.message.edit_text(
            f"❌ Xatolik yuz berdi:\n<code>{str(e)[:200]}</code>",
            parse_mode="HTML",
            reply_markup=back_kb("ap:excel_menu")
        )


# ─── /cancel ─────────────────────────────────────────────────────────────────

@router.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        return
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=main_panel_kb(message.from_user.id))


# ─── Yordamchi funksiyalar ────────────────────────────────────────────────────

def _post_preview(data: dict) -> str:
    msg_type = data.get("msg_type", "text")
    text = data.get("text", "")
    if msg_type == "text":
        return (text[:200] + "...") if len(text) > 200 else text
    return f"[{msg_type.upper()}] {(text[:100] + '...') if len(text) > 100 else text}"


async def _get_chat_ids(pool, target: str) -> list[int]:
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            if target == "groups":
                await cur.execute(
                    "SELECT chat_id FROM bot_chats WHERE chat_type IN ('group','supergroup')"
                )
            elif target == "channels":
                await cur.execute(
                    "SELECT chat_id FROM bot_chats WHERE chat_type = 'channel'"
                )
            elif target == "users":
                await cur.execute("SELECT id FROM users")
            else:
                await cur.execute("SELECT chat_id FROM bot_chats")
            return [r[0] for r in await cur.fetchall()]


async def _do_broadcast(bot, post_id: int, data: dict) -> tuple[int, int, int]:
    pool = await get_pool()
    target = data.get("target", "all")
    chat_ids = await _get_chat_ids(pool, target)

    msg_type = data.get("msg_type", "text")
    text = data.get("text", "")
    file_id = data.get("file_id")

    sent = fail = 0
    total_audience = 0

    for chat_id in chat_ids:
        try:
            if msg_type == "text":
                await bot.send_message(chat_id, text, parse_mode="HTML")
            elif msg_type == "photo":
                await bot.send_photo(chat_id, file_id, caption=text, parse_mode="HTML")
            elif msg_type == "video":
                await bot.send_video(chat_id, file_id, caption=text, parse_mode="HTML")
            elif msg_type == "document":
                await bot.send_document(chat_id, file_id, caption=text, parse_mode="HTML")
            elif msg_type == "animation":
                await bot.send_animation(chat_id, file_id, caption=text, parse_mode="HTML")
            sent += 1
            try:
                count = await bot.get_chat_member_count(chat_id)
                total_audience += count
            except Exception:
                total_audience += 1
        except Exception:
            fail += 1
        await asyncio.sleep(0.05)

    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "UPDATE scheduled_posts SET sent=TRUE, sent_count=%s, fail_count=%s, total_audience=%s WHERE id=%s",
                (sent, fail, total_audience, post_id)
            )

    return sent, fail, total_audience
